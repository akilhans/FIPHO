from datetime import date
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.core.cache import cache
from django.core.files.uploadedfile import SimpleUploadedFile
from django.conf import settings
from django.test import override_settings
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework.exceptions import ValidationError
from rest_framework.test import APITestCase

from registration.views import RegistrationAnonThrottle

from registration.models import (
    Country,
    Contestant,
    Delegation,
    DetailedRegistration,
    ParticipationRequest,
    Role,
    Subject,
    TeamLeader,
)
from registration.serializers import (
    ContestantSerializer,
    DelegationSerializer,
    DetailedRegistrationSerializer,
)


User = get_user_model()


def complete_fixed_roster(data):
    for group, required_count in (("team_leaders", 2), ("contestants", 4)):
        prefix = f"delegations[0][{group}][0]"
        original_items = [
            (key, value) for key, value in data.items() if key.startswith(prefix)
        ]
        for index in range(1, required_count):
            for key, value in original_items:
                copied_value = value
                if isinstance(value, SimpleUploadedFile):
                    value.seek(0)
                    copied_value = SimpleUploadedFile(
                        value.name,
                        value.read(),
                        value.content_type,
                    )
                    value.seek(0)
                data[key.replace(f"[{group}][0]", f"[{group}][{index}]")] = copied_value
    return data


class CountryAPITests(APITestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            username="admin", password="pass", is_staff=True
        )

    def test_country_is_public_read_admin_write(self):
        url = reverse("country-list")
        denied = self.client.post(url, {"name": "France"}, format="json")
        self.assertIn(denied.status_code, [401, 403])

        self.client.force_authenticate(user=self.admin)
        response = self.client.post(url, {"name": "France"}, format="json")
        self.assertEqual(response.status_code, 201)
        country_id = response.data["id"]

        dup_resp = self.client.post(url, {"name": "France"}, format="json")
        self.assertEqual(dup_resp.status_code, 400)

        self.client.force_authenticate(user=None)
        list_resp = self.client.get(url)
        self.assertEqual(list_resp.status_code, 200)
        self.assertEqual(len(list_resp.data), 1)

        self.client.force_authenticate(user=self.admin)
        detail_url = reverse("country-detail", args=[country_id])
        update_resp = self.client.put(detail_url, {"name": "New France"}, format="json")
        self.assertEqual(update_resp.status_code, 200)
        self.assertEqual(update_resp.data["name"], "New France")

        delete_resp = self.client.delete(detail_url)
        self.assertEqual(delete_resp.status_code, 204)


class ParticipationRequestAPITests(APITestCase):
    def setUp(self):
        self.country = Country.objects.create(name="Testland")
        self.role = Role.objects.create(name="Leader")
        self.subject = Subject.objects.create(name="Physics")
        self.url = reverse("participationrequest-list")
        self.admin = User.objects.create_user(
            username="admin", password="pass", is_staff=True
        )

    def test_public_can_create(self):
        data = {
            "full_name": "John",
            "country": self.country.id,
            "role": self.role.id,
            "subject": self.subject.id,
            "email": "john@example.com",
        }
        response = self.client.post(self.url, data, format="json")
        self.assertEqual(response.status_code, 201)

    @override_settings(FIPHO_MAX_STUDENTS=5, FIPHO_MAX_TEAM_LEADERS=2)
    def test_public_create_rejects_fipho_team_limits(self):
        base_data = {
            "full_name": "John",
            "country": self.country.id,
            "role": self.role.id,
            "subject": self.subject.id,
            "email": "john@example.com",
        }

        too_many_students = self.client.post(
            self.url,
            {**base_data, "number_of_students": 6},
            format="json",
        )
        self.assertEqual(too_many_students.status_code, 400)
        self.assertIn("number_of_students", too_many_students.data)

        too_many_leaders = self.client.post(
            self.url,
            {
                **base_data,
                "email": "leader-limit@example.com",
                "number_of_team_leaders": 3,
            },
            format="json",
        )
        self.assertEqual(too_many_leaders.status_code, 400)
        self.assertIn("number_of_team_leaders", too_many_leaders.data)

    def test_list_requires_admin(self):
        response = self.client.get(self.url)
        self.assertIn(response.status_code, [401, 403])

    def test_admin_can_list(self):
        self.client.force_authenticate(user=self.admin)
        data = {
            "full_name": "John",
            "country": self.country.id,
            "role": self.role.id,
            "subject": self.subject.id,
            "email": "john@example.com",
        }
        self.client.post(self.url, data, format="json")

        list_resp = self.client.get(self.url)
        self.assertEqual(list_resp.status_code, 200)
        self.assertEqual(len(list_resp.data["results"]), 1)


class DetailedRegistrationLimitTests(APITestCase):
    def test_second_step_requires_fixed_roster(self):
        serializer = DelegationSerializer()

        with self.assertRaisesMessage(
            ValidationError, "Exactly 4 contestants are required per delegation."
        ):
            serializer.validate_contestants([{}])

        with self.assertRaisesMessage(
            ValidationError, "Exactly 2 team leaders are required per delegation."
        ):
            serializer.validate_team_leaders([{}])

    def test_nested_replacement_requires_complete_people(self):
        country = Country.objects.create(name="Testland")
        registration = DetailedRegistration.objects.create(
            country=country,
            number_of_teams=1,
            confirm_information=True,
            agree_rules=True,
        )
        serializer = DetailedRegistrationSerializer(
            registration,
            data={
                "delegations": [{
                    "official_delegation_name": "Replacement",
                    "position": 1,
                    "team_leaders": [{}, {}],
                    "contestants": [{}, {}, {}, {}],
                }],
            },
            partial=True,
        )

        self.assertFalse(serializer.is_valid())
        self.assertIn("delegations", serializer.errors)

    def test_second_step_requires_people(self):
        country = Country.objects.create(name="Testland")
        serializer = DetailedRegistrationSerializer(
            data={
                "country": country.id,
                "confirm_information": True,
                "agree_rules": True,
            }
        )

        self.assertFalse(serializer.is_valid())
        self.assertIn("delegations", serializer.errors)

    def test_second_step_rejects_more_than_ten_teams(self):
        country = Country.objects.create(name="Testland")
        serializer = DetailedRegistrationSerializer(
            data={
                "country": country.id,
                "number_of_teams": 11,
                "confirm_information": True,
                "agree_rules": True,
            }
        )

        self.assertFalse(serializer.is_valid())
        self.assertIn("number_of_teams", serializer.errors)


class ContestantEligibilityTests(APITestCase):
    @staticmethod
    def _serializer():
        contestant = Contestant(
            full_name="Student",
            badge_name="Student",
            date_of_birth=date(2006, 5, 2),
            gender="Female",
            competition_subject="Physics",
            passport_number="AB1234567",
            passport_expiry_date=date(2030, 1, 1),
            t_shirt_size="M",
            food_type="Standard",
            passport_scan="passport.pdf",
            id_photo="photo.png",
            commitment_form="commitment.pdf",
            consent_form="consent.pdf",
        )
        return ContestantSerializer(instance=contestant)

    def test_rejects_contestant_born_on_cutoff(self):
        with self.assertRaises(ValidationError):
            self._serializer().validate({"date_of_birth": date(2006, 5, 1)})

    def test_accepts_contestant_born_after_cutoff(self):
        data = {"date_of_birth": date(2006, 5, 2)}
        self.assertEqual(self._serializer().validate(data), data)

    def test_rejects_contestant_born_in_2001(self):
        with self.assertRaises(ValidationError):
            self._serializer().validate({"date_of_birth": date(2001, 2, 2)})


class MediaAccessTests(APITestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            username="admin", password="pass", is_staff=True
        )

    def test_uploads_require_staff_access(self):
        with TemporaryDirectory() as tmpdir:
            media_root = Path(tmpdir)
            uploads_dir = media_root / "uploads" / "team_leaders" / "photos"
            uploads_dir.mkdir(parents=True)
            photo_path = uploads_dir / "photo.jpg"
            photo_path.write_bytes(b"photo data")

            with override_settings(MEDIA_ROOT=str(media_root)):
                anon_response = self.client.get("/media/uploads/team_leaders/photos/photo.jpg")
                self.assertEqual(anon_response.status_code, 403)

                self.client.force_login(self.admin)
                staff_response = self.client.get("/media/uploads/team_leaders/photos/photo.jpg")

        self.assertEqual(staff_response.status_code, 200)
        self.assertEqual(b"".join(staff_response.streaming_content), b"photo data")

    def test_non_public_media_requires_staff_access(self):
        with TemporaryDirectory() as tmpdir:
            media_root = Path(tmpdir)
            private_dir = media_root / "private"
            private_dir.mkdir()
            file_path = private_dir / "secret.txt"
            file_path.write_text("secret")

            with override_settings(MEDIA_ROOT=str(media_root)):
                response = self.client.get("/media/private/secret.txt")
                self.assertEqual(response.status_code, 403)

                self.client.force_login(self.admin)
                staff_response = self.client.get("/media/private/secret.txt")

        self.assertEqual(staff_response.status_code, 200)
        self.assertEqual(b"".join(staff_response.streaming_content), b"secret")

    def test_anon_html_request_redirects_to_admin_login(self):
        with TemporaryDirectory() as tmpdir:
            media_root = Path(tmpdir)
            uploads_dir = media_root / "uploads" / "team_leaders" / "photos"
            uploads_dir.mkdir(parents=True)
            (uploads_dir / "photo.jpg").write_bytes(b"photo data")

            with override_settings(MEDIA_ROOT=str(media_root)):
                response = self.client.get(
                    "/media/uploads/team_leaders/photos/photo.jpg",
                    HTTP_ACCEPT="text/html,application/xhtml+xml",
                )

        self.assertEqual(response.status_code, 302)
        location = response["Location"]
        self.assertTrue(
            location.startswith(settings.ADMIN_LOGIN_URL),
            f"Expected redirect to ADMIN_LOGIN_URL, got {location}",
        )
        self.assertIn("next=", location)

    def test_anon_non_html_request_returns_403(self):
        with TemporaryDirectory() as tmpdir:
            media_root = Path(tmpdir)
            uploads_dir = media_root / "uploads" / "team_leaders" / "photos"
            uploads_dir.mkdir(parents=True)
            (uploads_dir / "photo.jpg").write_bytes(b"photo data")

            with override_settings(MEDIA_ROOT=str(media_root)):
                response = self.client.get(
                    "/media/uploads/team_leaders/photos/photo.jpg",
                    HTTP_ACCEPT="application/json",
                )

        self.assertEqual(response.status_code, 403)


class FileValidationTests(APITestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            username="admin", password="pass", is_staff=True
        )
        self.country = Country.objects.create(name="Testland")
        self.registration = DetailedRegistration.objects.create(
            country=self.country,
            confirm_information=True,
            agree_rules=True,
        )
        self.delegation = Delegation.objects.create(
            registration=self.registration,
            official_delegation_name="Test Delegation",
            position=1,
        )
        self.client.force_authenticate(user=self.admin)

    def _make_file(self, name, content_type, size=100):
        return SimpleUploadedFile(name, b"x" * size, content_type=content_type)

    def test_valid_image_upload_accepted(self):
        url = reverse("teamleader-list")
        data = {
            "delegation": self.delegation.id,
            "full_name": "Leader",
            "badge_name": "Leader Badge",
            "date_of_birth": "1980-01-01",
            "gender": "Male",
            "passport_number": "TL123456",
            "email": "leader@example.com",
            "phone_number": "+1234567890",
            "role": "Team Leader",
            "t_shirt_size": "L",
            "food_type": "Standard",
            "id_photo": self._make_file("photo.jpg", "image/jpeg"),
            "passport_scan": self._make_file("scan.pdf", "application/pdf"),
            "consent_form": self._make_file("leader-consent.pdf", "application/pdf"),
        }
        response = self.client.post(url, data, format="multipart")
        self.assertEqual(response.status_code, 201)

    def test_invalid_file_type_rejected(self):
        url = reverse("teamleader-list")
        data = {
            "delegation": self.delegation.id,
            "full_name": "Leader",
            "badge_name": "Leader Badge",
            "date_of_birth": "1980-01-01",
            "gender": "Male",
            "passport_number": "TL123456",
            "email": "leader@example.com",
            "phone_number": "+1234567890",
            "role": "Team Leader",
            "t_shirt_size": "L",
            "food_type": "Standard",
            "id_photo": self._make_file("photo.exe", "application/x-msdownload"),
            "passport_scan": self._make_file("scan.pdf", "application/pdf"),
            "consent_form": self._make_file("leader-consent.pdf", "application/pdf"),
        }
        response = self.client.post(url, data, format="multipart")
        self.assertEqual(response.status_code, 400)
        self.assertIn("id_photo", response.data)

    def test_oversized_file_rejected(self):
        url = reverse("teamleader-list")
        oversized = self._make_file("photo.jpg", "image/jpeg", size=26 * 1024 * 1024)
        data = {
            "delegation": self.delegation.id,
            "full_name": "Leader",
            "badge_name": "Leader Badge",
            "date_of_birth": "1980-01-01",
            "gender": "Male",
            "passport_number": "TL123456",
            "email": "leader@example.com",
            "phone_number": "+1234567890",
            "role": "Team Leader",
            "t_shirt_size": "L",
            "food_type": "Standard",
            "id_photo": oversized,
            "passport_scan": self._make_file("scan.pdf", "application/pdf"),
            "consent_form": self._make_file("leader-consent.pdf", "application/pdf"),
        }
        response = self.client.post(url, data, format="multipart")
        self.assertEqual(response.status_code, 400)
        self.assertIn("id_photo", response.data)

    def test_required_fields_cannot_be_cleared(self):
        leader = TeamLeader.objects.create(
            delegation=self.delegation,
            full_name="Leader",
            badge_name="Leader Badge",
            date_of_birth="1980-01-01",
            gender="Male",
            passport_number="TL123456",
            email="leader@example.com",
            phone_number="+1234567890",
            role="Team Leader",
            t_shirt_size="L",
            food_type="Standard",
            passport_scan="leader-passport.pdf",
            id_photo="leader-photo.jpg",
            consent_form="leader-consent.pdf",
        )
        url = reverse("teamleader-detail", args=[leader.id])

        blank_name = self.client.patch(url, {"full_name": ""}, format="json")
        cleared_file = self.client.patch(
            url, {"passport_scan": None}, format="json"
        )

        self.assertEqual(blank_name.status_code, 400)
        self.assertIn("full_name", blank_name.data)
        self.assertEqual(cleared_file.status_code, 400)
        self.assertIn("passport_scan", cleared_file.data)

    def test_contestant_parental_consent_validates(self):
        url = reverse("contestant-list")
        data = {
            "delegation": self.delegation.id,
            "full_name": "Student",
            "badge_name": "Student Badge",
            "date_of_birth": "2011-01-01",
            "gender": "Male",
            "competition_subject": "Physics",
            "passport_number": "AB123",
            "passport_expiry_date": "2030-01-01",
            "t_shirt_size": "M",
            "food_type": "Standard",
            "passport_scan": self._make_file("scan.pdf", "application/pdf"),
            "id_photo": self._make_file("photo.jpg", "image/jpeg"),
            "commitment_form": self._make_file("commitment.pdf", "application/pdf"),
            "consent_form": self._make_file("consent.pdf", "application/pdf"),
            "parental_consent_form": self._make_file("consent.exe", "application/x-msdownload"),
        }
        response = self.client.post(url, data, format="multipart")
        self.assertEqual(response.status_code, 400)
        self.assertIn("parental_consent_form", response.data)


class RateLimitingTests(APITestCase):
    def setUp(self):
        self.country = Country.objects.create(name="Testland")
        self.role = Role.objects.create(name="Leader")
        self.subject = Subject.objects.create(name="Physics")
        self.url = reverse("participationrequest-list")

    @patch.object(RegistrationAnonThrottle, "THROTTLE_RATES", {"registration": "2/hour"})
    def test_rate_limiting(self):
        cache.clear()
        base_data = {
            "full_name": "John",
            "country": self.country.id,
            "role": self.role.id,
            "subject": self.subject.id,
        }
        r1 = self.client.post(self.url, {**base_data, "email": "a@example.com"}, format="json")
        self.assertEqual(r1.status_code, 201)
        r2 = self.client.post(self.url, {**base_data, "email": "b@example.com"}, format="json")
        self.assertEqual(r2.status_code, 201)
        r3 = self.client.post(self.url, {**base_data, "email": "c@example.com"}, format="json")
        self.assertEqual(r3.status_code, 429)


class ExportTests(APITestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            username="admin", password="pass", is_staff=True
        )
        self.country = Country.objects.create(name="Testland")
        self.role = Role.objects.create(name="Leader")
        self.subject = Subject.objects.create(name="Physics")

    def test_participation_export_requires_auth(self):
        url = reverse("export_participation_requests")
        response = self.client.get(url)
        self.assertIn(response.status_code, [401, 403])

    def test_participation_export_returns_xlsx(self):
        ParticipationRequest.objects.create(
            full_name="John",
            country=self.country,
            role=self.role,
            subject=self.subject,
            email="john@example.com",
        )
        self.client.force_authenticate(user=self.admin)
        url = reverse("export_participation_requests")
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.title, "Participation Requests")
        self.assertEqual(ws.cell(row=2, column=1).value, "John")

    def test_detailed_export_requires_auth(self):
        url = reverse("export_detailed_registrations")
        response = self.client.get(url)
        self.assertIn(response.status_code, [401, 403])

    def test_detailed_export_returns_xlsx_with_sheets(self):
        reg = DetailedRegistration.objects.create(
            country=self.country,
            confirm_information=True,
            agree_rules=True,
        )
        delegation = Delegation.objects.create(
            registration=reg,
            official_delegation_name="Test Team",
            position=1,
        )
        TeamLeader.objects.create(
            delegation=delegation,
            full_name="Leader One",
            email="leader@example.com",
            phone_number="+1234567890",
            role="Team Leader",
            passport_scan=SimpleUploadedFile(
                "leader-passport.pdf", b"x" * 10, "application/pdf"
            ),
        )
        Contestant.objects.create(
            delegation=delegation,
            full_name="Student One",
            date_of_birth="2008-05-15",
            competition_subject="Physics",
            passport_number="AB123",
            passport_expiry_date="2030-01-01",
            passport_scan=SimpleUploadedFile(
                "student-passport.pdf", b"x" * 10, "application/pdf"
            ),
        )
        self.client.force_authenticate(user=self.admin)
        url = reverse("export_detailed_registrations")
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        wb = load_workbook(BytesIO(response.content))
        self.assertEqual(
            wb.sheetnames,
            ["Registrations", "Delegations", "Team Leaders", "Contestants"],
        )

        ws_reg = wb["Registrations"]
        self.assertEqual(ws_reg.cell(row=2, column=4).value, "Test Team")
        reg_headers = [c.value for c in ws_reg[1]]
        self.assertIn("Confirm Information", reg_headers)
        self.assertIn("Agree Rules", reg_headers)
        self.assertEqual(
            ws_reg.cell(row=2, column=reg_headers.index("Confirm Information") + 1).value,
            "Yes",
        )
        self.assertEqual(
            ws_reg.cell(row=2, column=reg_headers.index("Agree Rules") + 1).value,
            "Yes",
        )

        ws_leaders = wb["Team Leaders"]
        self.assertEqual(ws_leaders.cell(row=2, column=4).value, "Leader One")
        leader_headers = [c.value for c in ws_leaders[1]]
        for header in ("Passport Scan", "ID Photo", "Consent Form"):
            self.assertIn(header, leader_headers)
        leader_passport = ws_leaders.cell(
            row=2, column=leader_headers.index("Passport Scan") + 1
        ).value
        self.assertIsInstance(leader_passport, str)
        self.assertTrue(leader_passport.startswith("http"))

        ws_contestants = wb["Contestants"]
        self.assertEqual(ws_contestants.cell(row=2, column=4).value, "Student One")
        contestant_headers = [c.value for c in ws_contestants[1]]
        for header in (
            "Passport Scan",
            "ID Photo",
            "Commitment Form",
            "Consent Form",
            "Parental Consent Form",
        ):
            self.assertIn(header, contestant_headers)
        contestant_passport = ws_contestants.cell(
            row=2, column=contestant_headers.index("Passport Scan") + 1
        ).value
        self.assertIsInstance(contestant_passport, str)
        self.assertTrue(contestant_passport.startswith("http"))


class PermissionsTests(APITestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            username="admin", password="pass", is_staff=True
        )
        self.country = Country.objects.create(name="Testland")
        self.role = Role.objects.create(name="Leader")
        self.subject = Subject.objects.create(name="Physics")

    def test_detailed_registration_requires_admin(self):
        url = reverse("detailed_registration_list")
        response = self.client.get(url)
        self.assertIn(response.status_code, [401, 403])

    def test_admin_detailed_registration_list_is_newest_first(self):
        older = DetailedRegistration.objects.create(
            country=self.country,
            confirm_information=True,
            agree_rules=True,
        )
        newer = DetailedRegistration.objects.create(
            country=self.country,
            confirm_information=True,
            agree_rules=True,
        )
        self.client.force_authenticate(user=self.admin)

        response = self.client.get(reverse("detailed_registration_list"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [item["id"] for item in response.data["results"]],
            [newer.id, older.id],
        )

    def test_detailed_registration_create_is_public(self):
        url = reverse("detailed_registration_list")
        data = {
            "country": str(self.country.id),
            "number_of_teams": "1",
            "confirm_information": "true",
            "agree_rules": "true",
            "delegations[0][official_delegation_name]": "Test Delegation",
            "delegations[0][position]": "1",
            "delegations[0][team_leaders][0][full_name]": "Leader One",
            "delegations[0][team_leaders][0][badge_name]": "L. One",
            "delegations[0][team_leaders][0][date_of_birth]": "1980-01-01",
            "delegations[0][team_leaders][0][gender]": "Male",
            "delegations[0][team_leaders][0][passport_number]": "TL123456",
            "delegations[0][team_leaders][0][email]": "leader@example.com",
            "delegations[0][team_leaders][0][phone_number]": "+1234567890",
            "delegations[0][team_leaders][0][role]": "Head of Delegation",
            "delegations[0][team_leaders][0][t_shirt_size]": "L",
            "delegations[0][team_leaders][0][food_type]": "Halal",
            "delegations[0][team_leaders][0][dietary_requirements]": "No peanuts",
            "delegations[0][team_leaders][0][passport_scan]": SimpleUploadedFile(
                "leader-passport.pdf", b"x" * 10, "application/pdf"
            ),
            "delegations[0][team_leaders][0][id_photo]": SimpleUploadedFile(
                "leader-photo.jpg", b"x" * 10, "image/jpeg"
            ),
            "delegations[0][team_leaders][0][consent_form]": SimpleUploadedFile(
                "leader-consent.pdf", b"x" * 10, "application/pdf"
            ),
            "delegations[0][contestants][0][full_name]": "Student One",
            "delegations[0][contestants][0][badge_name]": "S. One",
            "delegations[0][contestants][0][date_of_birth]": "2011-05-15",
            "delegations[0][contestants][0][gender]": "Female",
            "delegations[0][contestants][0][competition_subject]": "Physics",
            "delegations[0][contestants][0][passport_number]": "AB1234567",
            "delegations[0][contestants][0][passport_expiry_date]": "2030-01-01",
            "delegations[0][contestants][0][t_shirt_size]": "M",
            "delegations[0][contestants][0][food_type]": "Vegetarian",
            "delegations[0][contestants][0][dietary_requirements]": "No dairy",
            "delegations[0][contestants][0][special_requirements]": "",
            "delegations[0][contestants][0][passport_scan]": SimpleUploadedFile(
                "student-passport.pdf", b"x" * 10, "application/pdf"
            ),
            "delegations[0][contestants][0][id_photo]": SimpleUploadedFile(
                "student-photo.jpg", b"x" * 10, "image/jpeg"
            ),
            "delegations[0][contestants][0][commitment_form]": SimpleUploadedFile(
                "student-commitment.pdf", b"x" * 10, "application/pdf"
            ),
            "delegations[0][contestants][0][consent_form]": SimpleUploadedFile(
                "student-consent.pdf", b"x" * 10, "application/pdf"
            ),
        }
        response = self.client.post(
            url, complete_fixed_roster(data), format="multipart"
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(DetailedRegistration.objects.count(), 1)
        registration = DetailedRegistration.objects.get()
        self.assertEqual(registration.number_of_teams, 1)
        delegation = registration.delegations.get()
        self.assertEqual(delegation.official_delegation_name, "Test Delegation")
        self.assertEqual(delegation.team_leaders.count(), 2)
        self.assertEqual(delegation.contestants.count(), 4)
        self.assertEqual(delegation.team_leaders.first().badge_name, "L. One")
        self.assertEqual(delegation.team_leaders.first().food_type, "Halal")
        self.assertEqual(delegation.contestants.first().badge_name, "S. One")
        self.assertEqual(delegation.contestants.first().food_type, "Vegetarian")

    def test_detailed_registration_rejects_ineligible_contestant(self):
        url = reverse("detailed_registration_list")
        data = {
            "country": str(self.country.id),
            "number_of_teams": "1",
            "confirm_information": "true",
            "agree_rules": "true",
            "delegations[0][official_delegation_name]": "Test Delegation",
            "delegations[0][position]": "1",
            "delegations[0][team_leaders][0][full_name]": "Leader One",
            "delegations[0][team_leaders][0][badge_name]": "L. One",
            "delegations[0][team_leaders][0][date_of_birth]": "1980-01-01",
            "delegations[0][team_leaders][0][gender]": "Male",
            "delegations[0][team_leaders][0][passport_number]": "TL123456",
            "delegations[0][team_leaders][0][email]": "leader@example.com",
            "delegations[0][team_leaders][0][phone_number]": "+1234567890",
            "delegations[0][team_leaders][0][role]": "Head of Delegation",
            "delegations[0][team_leaders][0][t_shirt_size]": "L",
            "delegations[0][team_leaders][0][food_type]": "Halal",
            "delegations[0][team_leaders][0][passport_scan]": SimpleUploadedFile(
                "leader-passport.pdf", b"x" * 10, "application/pdf"
            ),
            "delegations[0][team_leaders][0][id_photo]": SimpleUploadedFile(
                "leader-photo.jpg", b"x" * 10, "image/jpeg"
            ),
            "delegations[0][team_leaders][0][consent_form]": SimpleUploadedFile(
                "leader-consent.pdf", b"x" * 10, "application/pdf"
            ),
            "delegations[0][contestants][0][full_name]": "Older Student",
            "delegations[0][contestants][0][badge_name]": "O. Student",
            "delegations[0][contestants][0][date_of_birth]": "2006-05-01",
            "delegations[0][contestants][0][gender]": "Female",
            "delegations[0][contestants][0][competition_subject]": "Physics",
            "delegations[0][contestants][0][passport_number]": "AB1234567",
            "delegations[0][contestants][0][passport_expiry_date]": "2030-01-01",
            "delegations[0][contestants][0][t_shirt_size]": "M",
            "delegations[0][contestants][0][food_type]": "Vegetarian",
            "delegations[0][contestants][0][passport_scan]": SimpleUploadedFile(
                "student-passport.pdf", b"x" * 10, "application/pdf"
            ),
            "delegations[0][contestants][0][id_photo]": SimpleUploadedFile(
                "student-photo.jpg", b"x" * 10, "image/jpeg"
            ),
            "delegations[0][contestants][0][commitment_form]": SimpleUploadedFile(
                "student-commitment.pdf", b"x" * 10, "application/pdf"
            ),
            "delegations[0][contestants][0][consent_form]": SimpleUploadedFile(
                "student-consent.pdf", b"x" * 10, "application/pdf"
            ),
        }
        response = self.client.post(
            url, complete_fixed_roster(data), format="multipart"
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("delegations", response.data)

    def test_team_leaders_requires_admin(self):
        url = reverse("teamleader-list")
        response = self.client.get(url)
        self.assertIn(response.status_code, [401, 403])

    def test_contestants_requires_admin(self):
        url = reverse("contestant-list")
        response = self.client.get(url)
        self.assertIn(response.status_code, [401, 403])

    def test_participation_request_create_is_public(self):
        url = reverse("participationrequest-list")
        data = {
            "full_name": "Public User",
            "country": self.country.id,
            "role": self.role.id,
            "subject": self.subject.id,
            "email": "public@example.com",
        }
        response = self.client.post(url, data, format="json")
        self.assertEqual(response.status_code, 201)

    def test_participation_request_list_requires_admin(self):
        url = reverse("participationrequest-list")
        response = self.client.get(url)
        self.assertIn(response.status_code, [401, 403])

        self.client.force_authenticate(user=self.admin)
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)


class CookieAuthTests(APITestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            username="admin", password="pass", is_staff=True
        )
        self.country = Country.objects.create(name="Testland")
        self.registration = DetailedRegistration.objects.create(
            country=self.country,
            confirm_information=True,
            agree_rules=True,
        )
        self.delegation = Delegation.objects.create(
            registration=self.registration,
            official_delegation_name="Test Delegation",
            position=1,
        )

    def _create_leader_with_passport(self):
        return TeamLeader.objects.create(
            delegation=self.delegation,
            full_name="Leader",
            email="leader@example.com",
            phone_number="+1234567890",
            role="Team Leader",
            passport_scan=SimpleUploadedFile(
                "passport.pdf", b"passport-data", "application/pdf"
            ),
        )

    def test_token_endpoint_sets_cookies(self):
        response = self.client.post(
            "/api/token/",
            {"username": "admin", "password": "pass"},
            format="json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("access_token", response.cookies)
        self.assertIn("refresh_token", response.cookies)
        self.assertTrue(response.cookies["access_token"]["httponly"])
        self.assertTrue(response.cookies["refresh_token"]["httponly"])

    def test_protected_media_with_cookie(self):
        with TemporaryDirectory() as tmpdir:
            with override_settings(MEDIA_ROOT=str(tmpdir), DEBUG=True):
                leader = self._create_leader_with_passport()
                login_resp = self.client.post(
                    "/api/token/",
                    {"username": "admin", "password": "pass"},
                    format="json",
                )
                self.assertEqual(login_resp.status_code, 200)

                response = self.client.get(f"/media/{leader.passport_scan.name}")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(b"".join(response.streaming_content), b"passport-data")

    def test_protected_media_without_auth(self):
        with TemporaryDirectory() as tmpdir:
            with override_settings(MEDIA_ROOT=str(tmpdir), DEBUG=True):
                leader = self._create_leader_with_passport()
                response = self.client.get(f"/media/{leader.passport_scan.name}")

        self.assertEqual(response.status_code, 403)

    def test_logout_clears_cookies(self):
        with TemporaryDirectory() as tmpdir:
            with override_settings(MEDIA_ROOT=str(tmpdir), DEBUG=True):
                leader = self._create_leader_with_passport()
                login_resp = self.client.post(
                    "/api/token/",
                    {"username": "admin", "password": "pass"},
                    format="json",
                )
                self.assertEqual(login_resp.status_code, 200)

                authed_resp = self.client.get(f"/media/{leader.passport_scan.name}")
                self.assertEqual(authed_resp.status_code, 200)

                logout_resp = self.client.post("/api/logout/")
                self.assertEqual(logout_resp.status_code, 204)

                response = self.client.get(f"/media/{leader.passport_scan.name}")

        self.assertEqual(response.status_code, 403)
