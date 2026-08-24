import logging
import os
import re
from pathlib import Path

import pytz
from django.conf import settings
from django.http import (
    FileResponse,
    Http404,
    HttpResponse,
    HttpResponseForbidden,
    HttpResponseRedirect,
    JsonResponse,
)
from drf_spectacular.utils import extend_schema
from openpyxl import Workbook
from rest_framework import generics, permissions, status, viewsets
from rest_framework.decorators import api_view, permission_classes
from rest_framework.exceptions import AuthenticationFailed
from rest_framework.response import Response
from rest_framework.throttling import AnonRateThrottle
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework_simplejwt.exceptions import InvalidToken, TokenError

logger = logging.getLogger("django.request")

from .models import (
    Country,
    Subject,
    Role,
    ParticipationRequest,
    DetailedRegistration,
    TeamLeader,
    Contestant,
)
from .serializers import (
    DETAILED_REGISTRATION_MAX_CONTESTANTS,
    DETAILED_REGISTRATION_MAX_TEAM_LEADERS,
    CountrySerializer,
    SubjectSerializer,
    RoleSerializer,
    ParticipationRequestSerializer,
    DetailedRegistrationSerializer,
    TeamLeaderSerializer,
    ContestantSerializer,
)


# ----------------- VIEWSETS FOR SIMPLE MODELS ----------------- #
class PublicReadAdminWriteViewSet(viewsets.ModelViewSet):
    def get_permissions(self):
        if self.action in {"list", "retrieve"}:
            return [permissions.AllowAny()]
        return [permissions.IsAdminUser()]


class CountryViewSet(PublicReadAdminWriteViewSet):
    queryset = Country.objects.all().order_by('name')
    serializer_class = CountrySerializer
    pagination_class = None


class SubjectViewSet(PublicReadAdminWriteViewSet):
    queryset = Subject.objects.all().order_by('name')
    serializer_class = SubjectSerializer
    pagination_class = None


class RoleViewSet(PublicReadAdminWriteViewSet):
    queryset = Role.objects.all().order_by('name')
    serializer_class = RoleSerializer
    pagination_class = None


class RegistrationAnonThrottle(AnonRateThrottle):
    scope = "registration"


class ParticipationRequestViewSet(viewsets.ModelViewSet):
    queryset = ParticipationRequest.objects.select_related('country', 'role', 'subject').all()
    serializer_class = ParticipationRequestSerializer

    def get_permissions(self):
        if self.action == 'create':
            return [permissions.AllowAny()]
        return [permissions.IsAdminUser()]

    def get_throttles(self):
        if self.action == "create":
            return [RegistrationAnonThrottle()]
        return []


class TeamLeaderViewSet(viewsets.ModelViewSet):
    queryset = TeamLeader.objects.select_related('delegation__registration').all()
    serializer_class = TeamLeaderSerializer
    permission_classes = [permissions.IsAdminUser]


class ContestantViewSet(viewsets.ModelViewSet):
    queryset = Contestant.objects.select_related('delegation__registration').all()
    serializer_class = ContestantSerializer
    permission_classes = [permissions.IsAdminUser]


# ----------------- DETAILED REGISTRATION VIEWS ----------------- #
class DetailedRegistrationListCreateView(generics.ListCreateAPIView):
    queryset = DetailedRegistration.objects.prefetch_related(
        'delegations__team_leaders', 'delegations__contestants'
    ).order_by('-created_at')
    serializer_class = DetailedRegistrationSerializer

    def get_permissions(self):
        if self.request.method == "POST":
            return [permissions.AllowAny()]
        return [permissions.IsAdminUser()]

    def get_throttles(self):
        if self.request.method == "POST":
            return [RegistrationAnonThrottle()]
        return []

    def create(self, request, *args, **kwargs):
        data = request.POST
        files = request.FILES

        total_upload_bytes = sum(
            uploaded_file.size
            for key in files
            for uploaded_file in files.getlist(key)
        )
        if total_upload_bytes > settings.FIPHO_MAX_TOTAL_UPLOAD_BYTES:
            return Response(
                {
                    "detail": (
                        "Total uploaded file size must not exceed "
                        f"{settings.FIPHO_MAX_TOTAL_UPLOAD_MB} MB."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        delegations = []

        def ensure_delegation(index):
            if index >= 10:
                raise ValueError("Maximum 10 delegation groups allowed")
            while len(delegations) <= index:
                delegations.append({"team_leaders": [], "contestants": []})
            return delegations[index]

        try:
            for source in (data, files):
                for key, value in source.items():
                    if not key.startswith("delegations["):
                        continue

                    delegation_match = re.fullmatch(
                        r"delegations\[(\d+)]\[(official_delegation_name|position)]",
                        key,
                    )
                    if delegation_match:
                        delegation_index = int(delegation_match.group(1))
                        delegation = ensure_delegation(delegation_index)
                        delegation[delegation_match.group(2)] = value
                        continue

                    person_match = re.fullmatch(
                        r"delegations\[(\d+)]\[(team_leaders|contestants)]\[(\d+)]\[([a-z_]+)]",
                        key,
                    )
                    if not person_match:
                        raise ValueError(f"Malformed nested field: {key}")

                    delegation_index = int(person_match.group(1))
                    collection_name = person_match.group(2)
                    person_index = int(person_match.group(3))
                    max_people = (
                        DETAILED_REGISTRATION_MAX_TEAM_LEADERS
                        if collection_name == "team_leaders"
                        else DETAILED_REGISTRATION_MAX_CONTESTANTS
                    )
                    if person_index >= max_people:
                        raise ValueError(
                            f"Too many {collection_name} in delegation {delegation_index + 1}"
                        )
                    delegation = ensure_delegation(delegation_index)
                    people = delegation[collection_name]
                    while len(people) <= person_index:
                        people.append({})
                    people[person_index][person_match.group(4)] = value
        except (ValueError, IndexError) as e:
            return Response(
                {"detail": f"Malformed form data: {e}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        main_data = {
            k: v for k, v in data.items()
            if not k.startswith("delegations[")
        }
        if delegations:
            main_data["delegations"] = delegations

        serializer = self.get_serializer(data=main_data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class DetailedRegistrationRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = DetailedRegistration.objects.prefetch_related(
        'delegations__team_leaders', 'delegations__contestants'
    ).order_by('-created_at')
    serializer_class = DetailedRegistrationSerializer
    permission_classes = [permissions.IsAdminUser]


# ----------------- UTILITY VIEW ----------------- #
PUBLIC_MEDIA_ROOTS = {"news", "media_library"}


def _auth_failure_response(request):
    wants_html = (
        request.method == "GET"
        and "text/html" in request.headers.get("Accept", "")
    )
    if wants_html:
        from urllib.parse import urlencode
        next_url = request.build_absolute_uri()
        qs = urlencode({"next": next_url})
        return HttpResponseRedirect(f"{settings.ADMIN_LOGIN_URL}?{qs}")
    return HttpResponseForbidden("Access denied")


def serve_protected_media(request, path):
    normalized_path = path.lstrip("/")
    requested_path = Path(normalized_path)

    if not requested_path.parts or any(part == ".." for part in requested_path.parts):
        raise Http404("Invalid file path")

    file_path = os.path.normpath(os.path.join(settings.MEDIA_ROOT, str(requested_path)))
    media_root = os.path.normpath(settings.MEDIA_ROOT)

    if not file_path.startswith(media_root):
        raise Http404("Invalid file path")

    is_public_resource = requested_path.parts and requested_path.parts[0] in PUBLIC_MEDIA_ROOTS

    if not is_public_resource:
        user = request.user

        if not user.is_authenticated:
            try:
                result = JWTAuthentication().authenticate(request)
                if result is not None:
                    user = result[0]
                    request.user = user
            except (AuthenticationFailed, InvalidToken, TokenError):
                pass

        if not user.is_authenticated:
            raw_cookie = request.COOKIES.get("access_token")
            if raw_cookie:
                try:
                    jwt_auth = JWTAuthentication()
                    validated_token = jwt_auth.get_validated_token(raw_cookie)
                    user = jwt_auth.get_user(validated_token)
                    request.user = user
                except (AuthenticationFailed, InvalidToken, TokenError):
                    pass

        if not user.is_authenticated or not user.is_active or not user.is_staff:
            return _auth_failure_response(request)

    if not os.path.exists(file_path):
        raise Http404("File does not exist")

    try:
        return FileResponse(open(file_path, "rb"))
    except OSError:
        raise Http404("File could not be read")


def _format_datetime(dt):
    if not dt:
        return ""
    tz = pytz.timezone("Asia/Tashkent")
    return dt.astimezone(tz).strftime("%d %b %Y, %H:%M")


@extend_schema(tags=["Export"], responses={(200, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"): bytes}, description="Export participation requests as XLSX")
@api_view(["GET"])
@permission_classes([permissions.IsAdminUser])
def export_participation_requests(request):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Participation Requests"

        headers = [
            "Full Name", "Country", "Role", "Subject", "Email", "WhatsApp",
            "Additional Number", "Students", "Team Leaders", "Created At",
        ]
        ws.append(headers)

        qs = ParticipationRequest.objects.select_related("country", "role", "subject").order_by("-created_at")
        for r in qs:
            ws.append([
                r.full_name,
                r.country.name,
                r.role.name,
                r.subject.name,
                r.email or "",
                r.whatsapp_number or "",
                r.additional_number or "",
                r.number_of_students,
                r.number_of_team_leaders,
                _format_datetime(r.created_at),
            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="participation-requests.xlsx"'
        wb.save(response)
        return response
    except Exception as e:
        logger.error("Failed to export participation requests: %s", e, exc_info=True)
        return JsonResponse({"detail": "Failed to generate export."}, status=500)


@extend_schema(tags=["Export"], responses={(200, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"): bytes}, description="Export detailed registrations as XLSX")
@api_view(["GET"])
@permission_classes([permissions.IsAdminUser])
def export_detailed_registrations(request):
    try:
        wb = Workbook()

        def _file_url(field):
            return request.build_absolute_uri(field.url) if field else ""

        # Sheet 1: Registrations
        ws_reg = wb.active
        ws_reg.title = "Registrations"
        ws_reg.append([
            "ID", "Country", "Participating Teams", "Delegation Names", "Team Leaders", "Contestants",
            "Created At", "Confirm Information", "Agree Rules",
        ])

        qs = DetailedRegistration.objects.select_related("country").prefetch_related(
            "delegations__team_leaders",
            "delegations__contestants",
        ).order_by("-created_at")
        for r in qs:
            delegations = list(r.delegations.all())
            ws_reg.append([
                r.id,
                r.country.name,
                r.number_of_teams,
                ", ".join(d.official_delegation_name for d in delegations),
                sum(len(d.team_leaders.all()) for d in delegations),
                sum(len(d.contestants.all()) for d in delegations),
                _format_datetime(r.created_at),
                "Yes" if r.confirm_information else "No",
                "Yes" if r.agree_rules else "No",
            ])

        # Sheet 2: Delegations
        ws_delegations = wb.create_sheet("Delegations")
        ws_delegations.append([
            "Registration ID",
            "Position",
            "Delegation Name",
            "Country",
            "Team Leaders",
            "Contestants",
        ])
        for r in qs:
            for delegation in r.delegations.all():
                ws_delegations.append([
                    r.id,
                    delegation.position,
                    delegation.official_delegation_name,
                    r.country.name,
                    len(delegation.team_leaders.all()),
                    len(delegation.contestants.all()),
                ])

        # Sheet 3: Team Leaders
        ws_leaders = wb.create_sheet("Team Leaders")
        ws_leaders.append([
            "Registration ID",
            "Delegation",
            "Country",
            "Full Name",
            "Badge Name",
            "Date of Birth",
            "Gender",
            "Passport Number",
            "Email",
            "Phone",
            "Role",
            "T-Shirt Size",
            "Food Type",
            "Dietary Requirements",
            "Passport Scan",
            "ID Photo",
            "Consent Form",
        ])
        for r in qs:
            for delegation in r.delegations.all():
                for leader in delegation.team_leaders.all():
                    ws_leaders.append([
                        r.id,
                        delegation.official_delegation_name,
                        r.country.name,
                        leader.full_name,
                        leader.badge_name or "",
                        str(leader.date_of_birth) if leader.date_of_birth else "",
                        leader.gender or "",
                        leader.passport_number or "",
                        leader.email,
                        leader.phone_number,
                        leader.role,
                        leader.t_shirt_size or "",
                        leader.food_type or "",
                        leader.dietary_requirements or "",
                        _file_url(leader.passport_scan),
                        _file_url(leader.id_photo),
                        _file_url(leader.consent_form),
                    ])

        # Sheet 4: Contestants
        ws_contestants = wb.create_sheet("Contestants")
        ws_contestants.append([
            "Registration ID",
            "Delegation",
            "Country",
            "Full Name",
            "Badge Name",
            "Date of Birth",
            "Gender",
            "Subject",
            "Passport Number",
            "Passport Expiry",
            "T-Shirt Size",
            "Food Type",
            "Dietary Requirements",
            "Special Requirements",
            "Passport Scan",
            "ID Photo",
            "Commitment Form",
            "Consent Form",
            "Parental Consent Form",
        ])
        for r in qs:
            for delegation in r.delegations.all():
                for contestant in delegation.contestants.all():
                    ws_contestants.append([
                        r.id, delegation.official_delegation_name, r.country.name, contestant.full_name,
                        contestant.badge_name or "",
                        str(contestant.date_of_birth) if contestant.date_of_birth else "",
                        contestant.gender or "", contestant.competition_subject, contestant.passport_number,
                        str(contestant.passport_expiry_date) if contestant.passport_expiry_date else "",
                        contestant.t_shirt_size or "", contestant.food_type or "", contestant.dietary_requirements or "", contestant.special_requirements or "",
                        _file_url(contestant.passport_scan),
                        _file_url(contestant.id_photo),
                        _file_url(contestant.commitment_form),
                        _file_url(contestant.consent_form),
                        _file_url(contestant.parental_consent_form),
                    ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="detailed-registrations.xlsx"'
        wb.save(response)
        return response
    except Exception as e:
        logger.error("Failed to export detailed registrations: %s", e, exc_info=True)
        return JsonResponse({"detail": "Failed to generate export."}, status=500)
